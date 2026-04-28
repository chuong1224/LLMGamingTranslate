"""
Excel format adapter for the generic translation system.

Handles Excel files (.xlsx, .xls) by:
1. Extracting text cells from all sheets
2. Grouping cells into translation blocks (indexed format)
3. Writing translations back while preserving structure and formatting
"""

import re
import io
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path

from .format_adapter import FormatAdapter
from .translation_unit import TranslationUnit


class ExcelAdapter(FormatAdapter):
    """
    Adapter for Excel files (.xlsx, .xls).

    Extracts string cells from all worksheets, groups them into blocks,
    translates each block using [N] indexed format, and writes the
    translations back to a new workbook preserving all formatting.
    """

    CELLS_PER_BLOCK = 15

    def __init__(self, input_file_path: str, output_file_path: str, config: Dict[str, Any]):
        super().__init__(input_file_path, output_file_path, config)
        self.cells: List[Dict] = []
        self.blocks: List[List[int]] = []
        self.translations: Dict[int, str] = {}

    async def prepare_for_translation(self) -> bool:
        """Read workbook and collect all translatable string cells."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(self.input_file_path, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.strip():
                            self.cells.append({
                                'sheet': sheet_name,
                                'row': cell.row,
                                'col': cell.column,
                                'original_text': cell.value,
                            })

            if not self.cells:
                return False

            cells_per_block = self.config.get('cells_per_block', self.CELLS_PER_BLOCK)
            self.blocks = [
                list(range(i, min(i + cells_per_block, len(self.cells))))
                for i in range(0, len(self.cells), cells_per_block)
            ]
            return True

        except Exception:
            return False

    def get_translation_units(self) -> List[TranslationUnit]:
        """Create one TranslationUnit per block of cells."""
        units = []

        for block_idx, cell_indices in enumerate(self.blocks):
            lines = []
            for local_idx, global_idx in enumerate(cell_indices):
                text = self.cells[global_idx]['original_text']
                lines.append(f"[{local_idx + 1}] {text}")

            context_before = ""
            context_after = ""
            if block_idx > 0:
                prev_last = self.blocks[block_idx - 1][-1]
                context_before = self.cells[prev_last]['original_text']
            if block_idx < len(self.blocks) - 1:
                next_first = self.blocks[block_idx + 1][0]
                context_after = self.cells[next_first]['original_text']

            unit = TranslationUnit(
                unit_id=f"block_{block_idx}",
                content="\n".join(lines),
                context_before=context_before,
                context_after=context_after,
                metadata={
                    'block_index': block_idx,
                    'cell_indices': cell_indices,
                },
            )
            units.append(unit)

        return units

    async def save_unit_translation(self, unit_id: str, translated_content: str) -> bool:
        """Parse [N] indexed response and store per-cell translations."""
        try:
            block_idx = int(unit_id.split('_')[1])
            if block_idx >= len(self.blocks):
                return False

            cell_indices = self.blocks[block_idx]
            parsed = self._parse_indexed_response(translated_content, len(cell_indices))

            for local_idx, global_idx in enumerate(cell_indices):
                translation = parsed.get(local_idx + 1, '').strip()
                self.translations[global_idx] = translation or self.cells[global_idx]['original_text']

            return True

        except Exception:
            return False

    def _parse_indexed_response(self, text: str, expected: int) -> Dict[int, str]:
        """
        Parse LLM response that uses [N] markers.

        Handles both clean format ([1] text\\n[2] text) and cases where
        the model omits some markers or adds extra text.
        """
        result: Dict[int, str] = {}

        pattern = re.compile(r'\[(\d+)\]\s*(.*?)(?=\n?\[\d+\]|\Z)', re.DOTALL)
        for m in pattern.finditer(text):
            idx = int(m.group(1))
            content = m.group(2).strip()
            if 1 <= idx <= expected:
                result[idx] = content

        return result

    async def reconstruct_output(self, bilingual: bool = False) -> bytes:
        """Write translated cells back into the workbook and return as bytes."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(self.input_file_path)

            for cell_idx, cell_info in enumerate(self.cells):
                translated = self.translations.get(cell_idx, cell_info['original_text'])
                sheet_name = cell_info['sheet']
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                target_cell = ws.cell(row=cell_info['row'], column=cell_info['col'])
                if bilingual:
                    target_cell.value = f"{cell_info['original_text']}\n{translated}"
                else:
                    target_cell.value = translated

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        except Exception:
            with open(self.input_file_path, 'rb') as f:
                return f.read()

    async def resume_from_checkpoint(self, checkpoint_data: Dict[str, Any]) -> int:
        """Restore per-cell translations from checkpoint."""
        try:
            for chunk_data in checkpoint_data.get('chunks', []):
                if chunk_data.get('status') != 'completed':
                    continue
                metadata = chunk_data.get('chunk_data', {})
                translated_text = chunk_data.get('translated_text', '')
                cell_indices = metadata.get('cell_indices', [])
                if not (translated_text and cell_indices):
                    continue
                parsed = self._parse_indexed_response(translated_text, len(cell_indices))
                for local_idx, global_idx in enumerate(cell_indices):
                    translation = parsed.get(local_idx + 1, '').strip()
                    if translation:
                        self.translations[global_idx] = translation

            return checkpoint_data.get('resume_from_index', 0)

        except Exception:
            return 0

    async def cleanup(self):
        """No temporary files to clean up."""
        pass

    @property
    def format_name(self) -> str:
        return "xlsx"
